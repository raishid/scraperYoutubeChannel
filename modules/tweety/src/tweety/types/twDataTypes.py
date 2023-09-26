import base64
import datetime
import json
import os.path
import re
import html
import time
import traceback
import warnings
from typing import Callable
from dateutil import parser
import openpyxl
import dateutil
from ..exceptions_ import UserNotFound, UserProtected
from ..utils import *


def deprecated(func):
    """

    This is a decorator which can be used to mark functions
    as deprecated. It will result in a warning being emitted
    when the function is used.

    """

    def new_func(*args, **kwargs):
        warnings.warn("Call to deprecated function {}.".format(func.__name__), category=DeprecationWarning)
        return func(*args, **kwargs)

    new_func.__name__ = func.__name__
    new_func.__doc__ = func.__doc__
    new_func.__dict__.update(func.__dict__)
    return new_func


class Excel:
    def __init__(self, tweets, filename=None, append=False):
        self.tweets = tweets
        self.filename = filename
        self.author = self._get_author()
        self._append = append
        self._get_sheet()
        self.max_row = self._get_max_row()
        self._write_data()

    def _get_sheet(self):
        if self._append and self.filename:
            self.workbook = openpyxl.load_workbook(self.filename)
            self.worksheet = self.workbook.active
        else:
            self._append = False
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.create_sheet("tweets")

    def _get_author(self):
        for tweet in self.tweets:
            if hasattr(tweet, "author"):
                return tweet.author.username

        return ""

    def _get_max_row(self):
        if self._append:
            for row in reversed(self.worksheet.iter_rows(values_only=True)):
                if any(cell for cell in row):
                    return row[0].row

        self._set_headers()
        return 1

    def _set_headers(self):
        for index, value in enumerate(WORKBOOK_HEADERS, start=1):
            self.worksheet.cell(row=1, column=index).value = value

    def _write_tweet(self, tweet):
        self.worksheet[f'A{self.max_row + 1}'] = tweet.date
        self.worksheet[f'B{self.max_row + 1}'] = tweet.author.name
        self.worksheet[f'C{self.max_row + 1}'] = tweet.id
        self.worksheet[f'D{self.max_row + 1}'] = tweet.text
        self.worksheet[f'E{self.max_row + 1}'] = tweet.is_retweet
        self.worksheet[f'F{self.max_row + 1}'] = tweet.is_reply
        self.worksheet[f'G{self.max_row + 1}'] = tweet.language
        self.worksheet[f'H{self.max_row + 1}'] = tweet.likes
        self.worksheet[f'I{self.max_row + 1}'] = tweet.retweet_counts
        self.worksheet[f'J{self.max_row + 1}'] = tweet.source
        self.worksheet[f'K{self.max_row + 1}'] = ",".join(
            [media.expanded_url for media in tweet.media]) if tweet.media else ""
        self.worksheet[f'L{self.max_row + 1}'] = ",".join(
            [user_mention.screen_name for user_mention in tweet.user_mentions]) if tweet.user_mentions else ""
        self.worksheet[f'M{self.max_row + 1}'] = ",".join(
            [url['expanded_url'] for url in tweet.urls]) if tweet.urls else ""
        self.worksheet[f'N{self.max_row + 1}'] = ",".join(
            [hashtag['text'] for hashtag in tweet.hashtags]) if tweet.hashtags else ""
        self.worksheet[f'O{self.max_row + 1}'] = ",".join([symbol for symbol in tweet.symbols]) if tweet.symbols else ""
        self.max_row += 1

    def _write_data(self):
        for tweet in self.tweets:
            if isinstance(tweet, Tweet):
                self._write_tweet(tweet)
            elif isinstance(tweet, SelfThread):
                for _threadedTweet in tweet:
                    self._write_tweet(_threadedTweet)

        if not self.filename:
            self.filename = f"tweets-{self.author}.xlsx"

        try:
            self.workbook.remove("sheet")
        except ValueError:
            pass

        self.workbook.save(self.filename)


class Tweet(dict):
    def __init__(self, tweet, client, full_http_response=None):  # noqa
        super().__init__()

        self._comments_cursor = None
        self._raw = tweet
        self._client = client
        self._full_http_response = full_http_response
        self._format_tweet()

        for key, value in vars(self).items():
            if not str(key).startswith("_"):
                self[key] = value

    def __eq__(self, other):
        if isinstance(other, Tweet):
            return str(self.id) == str(other.id) and str(self.author.id) == str(other.author.id)
        elif isinstance(other, (int, str)):
            return str(self.id) == str(other)

        return False

    def __repr__(self):
        return f"Tweet(id={self.id}, author={self.author}, created_on={self.created_on})"  # noqa

    def __iter__(self):
        if self.threads:  # noqa
            for thread in self.threads:  # noqa
                yield thread

    def _format_tweet(self):
        self._tweet = find_objects(self._raw, "__typename", ["Tweet", "TweetWithVisibilityResults"], recursive=False)

        if self._tweet.get('tweet'):
            self._tweet = self._tweet['tweet']

        self._card = self._tweet.get('card')
        self.original_tweet = self._get_original_tweet()
        self.id = self._get_id()
        self.created_on = self.date = self._get_date()
        self.author = self._get_author()
        self.is_retweet = self._is_retweet()
        self.retweeted_tweet = self._get_retweeted_tweet()
        self.rich_text = self._get_rich_text()
        self.text = self.tweet_body = self._get_tweet_text()
        self.is_quoted = self._is_quoted()
        self.quoted_tweet = self._get_quoted_tweet()
        self.is_reply = self._is_reply()
        self.is_sensitive = self._is_sensitive()
        self.reply_counts = self._get_reply_counts()
        self.quote_counts = self._get_quote_counts()
        self.replied_to = None
        self.bookmark_count = self._get_bookmark_count()
        self.vibe = self._get_vibe()
        self.views = self._get_views()
        self.language = self._get_language()
        self.likes = self._get_likes()
        self.place = self._get_place()
        self.retweet_counts = self._get_retweet_counts()
        self.source = self._get_source()
        self.audio_space_id = self._get_audio_space()
        self.voice_info = None  # TODO
        self.media = self._get_tweet_media()
        self.pool = self._get_pool()
        self.user_mentions = self._get_tweet_mentions()
        self.urls = self._get_tweet_urls()
        self.has_moderated_replies = self._get_has_moderated_replies()
        self.hashtags = self._get_tweet_hashtags()
        self.symbols = self._get_tweet_symbols()
        self.community_note = self._get_community_note()
        self.community = self._get_community()
        self.url = self._get_url()
        self.threads = self.get_threads()
        self.comments = []

    def _get_url(self):
        return "https://twitter.com/{}/status/{}".format(
            self.author.username, self.id
        )

    def like(self):
        return self._client.like_tweet(self.id)

    def retweet(self):
        return self._client.retweet_tweet(self.id)

    def _get_original_tweet(self):
        tweet = self._tweet

        if tweet.get('tweet'):
            tweet = tweet['tweet']

        return tweet['legacy'] if tweet.get('legacy') else tweet

    def _get_has_moderated_replies(self):
        return self._tweet.get('hasModeratedReplies', False)

    def get_threads(self):
        _threads = []
        if not self._full_http_response:
            return _threads

        instruction = find_objects(self._full_http_response, "type", "TimelineAddEntries")
        if not instruction:
            return _threads

        entries = instruction.get('entries', [])
        for entry in entries:

            if str(entry['entryId'].split("-")[0]) == "conversationthread":
                _thread = [i for i in entry['content']['items']]
                self_threads = [i for i in _thread if i['item']['itemContent'].get('tweetDisplayType') == "SelfThread"]

                if len(self_threads) == 0:
                    continue

                for _ in self_threads:
                    try:
                        parsed = Tweet(_, self._client, None)
                        _threads.append(parsed)
                    except:
                        pass

            elif str(entry['entryId'].split("-")[0]) == "tweet" and entry['content']['itemContent']['tweetDisplayType'] == "SelfThread":
                try:
                    parsed = Tweet(entry, self._client, None)
                    _threads.append(parsed)
                except:
                    pass
        return _threads

    def get_comments(self, pages=1, wait_time=2, cursor=None):

        list(self.iter_comments(pages, wait_time, cursor))
        return self.comments

    def iter_comments(self, pages=1, wait_time=2, cursor=None):
        if not wait_time:
            wait_time = 0

        self._comments_cursor = cursor if cursor else self._comments_cursor
        pages = pages
        for page in range(1, int(pages) + 1):
            _, comments, self._comments_cursor = self._get_comments(self._comments_cursor, self._full_http_response)

            yield self, comments

            if cursor != self._comments_cursor and page != pages:
                time.sleep(wait_time)
            else:
                break

    def _get_comments(self, cursor=None, response=None):
        _comments = []
        _cursor = cursor

        if not response:
            response = self._client.http.get_tweet_detail(self.id, _cursor)

        self._full_http_response = None

        instruction = find_objects(response, "type", "TimelineAddEntries")
        for entry in instruction['entries']:
            if str(entry['entryId'].split("-")[0]) == "conversationthread":
                thread = [i for i in entry['content']['items']]

                if len(thread) == 0:
                    continue

                try:
                    parsed = ConversationThread(self, thread, self._client)
                    _comments.append(parsed)
                    self.comments.append(parsed)
                except:
                    pass

            elif "cursor-bottom" in str(entry['entryId']) or "cursor-showmorethreads" in str(entry['entryId']):
                _cursor = entry['content']['itemContent']['value']
        return self, _comments, _cursor

    def _get_pool(self):
        if not self._card or "poll" not in self._card.get('legacy', {}).get('name', ''):
            return None

        return Poll(self._card)

    def _get_audio_space(self):
        if not self._card or "audiospace" not in self._card.get('legacy', {}).get('name', ''):
            return None

        for value in self._card.get('legacy', {}).get('binding_values', []):
            if value['key'] == "id":
                return value['value']['string_value']

        return None

    def _get_community_note(self):
        if self._tweet.get("birdwatch_pivot"):
            return self._tweet['birdwatch_pivot']['subtitle']['text']

        return None

    def _get_date(self):
        date = self.original_tweet.get("created_at")

        if date:
            return dateutil.parser.parse(date)

        return None

    def _get_id(self):
        return self._tweet.get('rest_id')

    def _get_community(self):
        return Community(self._tweet['community_results'], self._client) if self._tweet.get('community_results') else None

    def _get_author(self):
        if self._tweet.get("core"):
            return User(self._tweet['core'], self._client)

        if self._tweet.get("author"):
            return User(self._tweet['author'], self._client)

        return None

    def _get_retweeted_tweet(self):
        if self.is_retweet and self.original_tweet.get("retweeted_status_result"):
            retweet = self.original_tweet['retweeted_status_result']['result']
            return Tweet(retweet, self._client)

        return None

    def _get_quoted_tweet(self):
        if not self.is_quoted:
            return None

        try:
            if self._tweet.get("quoted_status_result"):
                raw_tweet = self._tweet['quoted_status_result']['result']
                return Tweet(raw_tweet, self._client)

            if self.original_tweet.get('retweeted_status_result'):
                raw_tweet = self.original_tweet['retweeted_status_result']['result']['quoted_status_result']['result']
                return Tweet(raw_tweet, self._client)

            return None
        except:
            return None

    def _get_vibe(self):
        if self._tweet.get("vibe"):
            vibeImage = self._tweet['vibe']['imgDescription']
            vibeText = self._tweet['vibe']['text']
            return f"{vibeImage} {vibeText}"

        return ""

    def _get_views(self):
        if self._tweet.get("views"):
            return self._tweet['views'].get('count', 'Unavailable')

        return 0

    def get_reply_to(self):
        if not self.is_reply:
            return None

        tweet_id = self.original_tweet['in_reply_to_status_id_str']
        if not self._full_http_response:
            response = self._client.http.get_tweet_detail(tweet_id)
        else:
            response = self._full_http_response

        try:
            if response['data'].get('threaded_conversation_with_injections_v2'):
                entries = response['data']['threaded_conversation_with_injections_v2']['instructions'][0]['entries']
            else:
                entries = response['data']['search_by_raw_query']['search_timeline']['timeline']['instructions'][0][
                    'entries']

            for entry in entries:
                if str(entry['entryId']).split("-")[0] == "tweet" and str(
                        entry['content']['itemContent']['tweet_results']['result']['rest_id']) == str(tweet_id):
                    raw_tweet = entry['content']['itemContent']['tweet_results']['result']
                    return Tweet(raw_tweet, self._client.http)
        except:
            pass

        return None

    def _is_sensitive(self):
        return self.original_tweet.get("possibly_sensitive", False)

    def _get_reply_counts(self):
        return self.original_tweet.get("reply_count", 0)

    def _get_quote_counts(self):
        return self.original_tweet.get("quote_count", 0)

    def _is_retweet(self):
        if self.original_tweet.get('retweeted'):
            return self.original_tweet['retweeted']

        if str(self.original_tweet.get('full_text', "")).startswith("RT"):
            return True

        return False

    def _is_reply(self):
        tweet_keys = list(self.original_tweet.keys())
        required_keys = ["in_reply_to_status_id_str", "in_reply_to_user_id_str", "in_reply_to_screen_name"]
        return all(x in tweet_keys for x in required_keys)

    def _is_quoted(self):
        if self.original_tweet.get("is_quote_status"):
            return True

        return False

    def _get_language(self):
        return self.original_tweet.get('lang', "")

    def _get_likes(self):
        return self.original_tweet.get("favorite_count", 0)

    def _get_place(self):
        if self.original_tweet.get('place'):
            return Place(self.original_tweet['place'])

        return None

    def _get_retweet_counts(self):
        return self.original_tweet.get('retweet_count', 0)

    def _get_source(self):
        if self._tweet.get('source'):
            return str(self._tweet['source']).split(">")[1].split("<")[0]

        return ""

    def _get_rich_text(self):
        note_tweet = self._tweet.get('note_tweet')

        if not note_tweet:
            return None

        return RichText(note_tweet, self)

    def _get_tweet_text(self):
        if self.is_retweet and self.original_tweet.get("retweeted_status_result"):
            return self.retweeted_tweet.text

        if self.rich_text:
            return self.rich_text.text

        if self.original_tweet.get('full_text'):
            return html.unescape(self.original_tweet['full_text'])

        return ""

    def _get_tweet_media(self):
        return [Media(media, self._client) for media in self.original_tweet.get("extended_entities", {}).get("media", [])]

    def _get_tweet_mentions(self):
        users = [ShortUser(user) for user in self.original_tweet.get("entities", {}).get("user_mentions", [])]

        if self.rich_text:
            for user in self.rich_text.user_mentions:
                if user not in users:
                    users.append(user)
        return users

    def _get_bookmark_count(self):
        return self.original_tweet.get("bookmark_count", None)

    def _get_tweet_urls(self):
        if not self.original_tweet.get("entities"):
            return []

        if not self.original_tweet['entities'].get("urls"):
            return []

        return [url for url in self.original_tweet['entities']['urls']]

    def _get_tweet_hashtags(self):
        if not self.original_tweet.get("entities"):
            return []

        if not self.original_tweet['entities'].get("hashtags"):
            return []

        return [hashtag for hashtag in self.original_tweet['entities']['hashtags']]

    def _get_tweet_symbols(self):
        if not self.original_tweet.get("entities"):
            return []

        if not self.original_tweet['entities'].get("symbols"):
            return []

        return [symbol for symbol in self.original_tweet['entities']['symbols']]


class RichText(dict):
    HTML_TAGS = {
        "Bold": "b",
        "Italic": "em"
    }

    def __init__(self, data, tweet):
        super().__init__()
        self._raw = data
        self._tweet = tweet
        self._note = self._raw['note_tweet_results']['result']
        self._entities = self._note['entity_set']
        self.tags = self._get_tags()
        self.id = self._get_id()
        self.text = self._get_text()
        self.hashtags = self._get_hashtags()
        self.urls = self._get_urls()
        self.symbols = self._get_symbols()
        self.user_mentions = self._get_mentions()
        self.media = self._get_media()

        for key, value in vars(self).items():
            if not str(key).startswith("_"):
                self[key] = value

    def __repr__(self):
        return "RichText(id={})".format(self.id)

    def __eq__(self, other):
        if isinstance(other, RichText):
            return self.id == other.id
        elif isinstance(other, str):
            return str(self.id) == str(other)

        return False

    def _get_id(self):
        return self._note.get('id')

    def _get_text(self):
        return self._note.get('text', '')

    def _get_hashtags(self):
        return self._entities.get('hashtags', [])

    def _get_urls(self):
        return self._entities.get('urls', [])

    def _get_symbols(self):
        return self._entities.get('symbols', [])

    def _get_tags(self):
        return [RichTag(i) for i in self._note.get('richtext', {}).get('richtext_tags', [])]

    def _get_media(self):
        return [i for i in self._note.get('media', {}).get('inline_media', [])]

    def _get_mentions(self):
        return [ShortUser(i) for i in self._entities.get('user_mentions', [])]

    def get_html(self):
        tags = self.tags
        tags.extend(self.user_mentions)
        tags.extend(self.media)
        ordered_tags = sorted(tags, key=lambda x: x.from_index if hasattr(x, "from_index") else x['index'], reverse=True)
        thisHtml = self.text

        for tag in ordered_tags:
            text = self.text[tag.from_index:tag.to_index] if hasattr(tag, "from_index") else self.text[tag['index']]
            if isinstance(tag, RichTag) and hasattr(tag, "types"):
                for _type in tag.types:
                    tag_name = self.HTML_TAGS.get(_type)
                    new_text = "<{tag_name}>{text}</{tag_name}>".format(tag_name=tag_name, text=text)
                    thisHtml = replace_between_indexes(thisHtml, tag.from_index, tag.to_index, new_text)
            elif isinstance(tag, ShortUser):
                new_text = "<a href='{}'>@{}</a>".format(tag.url, tag.username)
                thisHtml = replace_between_indexes(thisHtml, tag.from_index, tag.to_index, new_text)
            elif tag.get('media_id'):
                for media in self._tweet.media:
                    if media == tag['media_id']:
                        new_text = "<img src='{}'><br>".format(media.direct_url)
                        thisHtml = replace_between_indexes(thisHtml, tag['index'], tag['index'], new_text)
                        break

        return f"<pre>{thisHtml}</pre>"


class RichTag(dict):
    def __init__(self, data):
        super().__init__()
        self._raw = data
        self.from_index = self._raw.get('from_index')
        self.to_index = self._raw.get('to_index')
        self.types = self._raw.get('richtext_types', [])

        for key, value in vars(self).items():
            if not str(key).startswith("_"):
                self[key] = value

    def __repr__(self):
        return "RichTag(from_index={}, to_index={}, types={})".format(
            self.from_index, self.to_index, self.types
        )

class SelfThread(dict):
    def __init__(self, conversation_tweet, client, full_response):
        super().__init__()
        self._client = client
        self._raw = conversation_tweet
        self.tweets = []
        self.all_tweets_id = self._get_all_tweet_ids()
        self._format_tweet()

        for key, value in vars(self).items():
            if not str(key).startswith("_"):
                self[key] = value

    def _format_tweet(self):
        for item in self._raw['content']['items']:
            try:
                parsed = Tweet(item, self._client, None)
                self.tweets.append(parsed)
            except:
                pass

    def __iter__(self):
        if self.tweets:
            for tweet in self.tweets:  # noqa
                yield tweet

    def expand(self):
        tweet = self._client.tweet_detail(self.tweets[0].id)
        self.tweets = tweet.threads

    def _get_all_tweet_ids(self):
        return self._raw['content']['metadata']['conversationMetadata']['allTweetIds']

    def __repr__(self):
        return "SelfThread(tweets={}, all_tweets={})".format(
            len(self.tweets), len(self.all_tweets_id)
        )


class ConversationThread(dict):
    def __init__(self, parent_tweet, thread_tweets, client):
        super().__init__()
        self._client = client
        self.tweets = []
        self.parent = parent_tweet
        self.cursor = None
        self._threads = thread_tweets
        self._format_threads()

        for key, value in vars(self).items():
            if not str(key).startswith("_"):
                self[key] = value

    def __repr__(self):
        return "ConversationThread(parent={}, tweets={})".format(
            self.parent, len(self.tweets)
        )

    def _format_threads(self):
        for thread in self._threads:
            entry_type = str(thread['entryId']).split("-")[-2].lower()
            if entry_type == "tweet":
                self.tweets.append(Tweet(thread, self._client, None))
            elif entry_type == "showmore":
                self.cursor = thread['item']['itemContent']['value']

    def expand(self):
        if not self.cursor:
            return self.tweets

        response = self._client.http.get_tweet_detail(self.parent.id, self.cursor)
        moduleItems = find_objects(response, "moduleItems", None)

        if not moduleItems or len(moduleItems) == 0:
            return self.tweets

        for item in moduleItems:
            tweet = find_objects(item, "__typename", ["Tweet", "TweetWithVisibilityResults"], recursive=False)
            if tweet:
                self.tweets.append(Tweet(tweet, self._client, None))

        return self.tweets


class Media(dict):
    def __init__(self, media_dict, client):
        super().__init__()
        self._raw = media_dict
        self._client = client
        self.display_url = self._raw.get("display_url")
        self.expanded_url = self._raw.get("expanded_url")
        self.id = self._raw.get("id_str")
        self.alt_text = self._raw.get("ext_alt_text")
        self.indices = self._raw.get("indices")
        self.media_url_https = self.direct_url = self._get_direct_url()
        self.type = self._raw.get("type")
        self.url = self._raw.get("url")
        self.features = self._raw.get("features")
        self.media_key = self._raw.get("media_key")
        self.mediaStats = self._raw.get("mediaStats")
        self.sizes = [MediaSize(k, v) for k, v in self._raw.get("sizes", {}).items() if self._raw.get('sizes')]
        self.original_info = self._raw.get("original_info")
        self.file_format = self._get_file_format()
        self.streams = []

        if self.type == "video" or self.type == "animated_gif":
            self._parse_video_streams()

        for k, v in vars(self).items():
            if not k.startswith("_"):
                self[k] = v

    def __eq__(self, other):
        if isinstance(other, Media):
            return self.id == other.id
        elif isinstance(other, (int, str)):
            return str(self.id) == str(other)

        return False

    def _get_direct_url(self):
        url = self._raw.get("media_url_https")

        return url

    def _get_file_format(self):
        filename = os.path.basename(self.media_url_https).split("?")[0]
        return filename.split(".")[-1] if self.type == "photo" else "mp4"

    def _parse_video_streams(self):
        videoDict = self._raw.get("video_info")

        if not videoDict:
            return

        for i in videoDict.get("variants"):
            if not i.get("content_type").split("/")[-1] == "x-mpegURL":
                self.streams.append(
                    Stream(i, videoDict.get("duration_millis", 0), videoDict.get("aspect_ratio"), self._client))

    def __repr__(self):
        return f"Media(id={self.id}, type={self.type})"

    def download(self, filename: str = None, progress_callback: Callable[[str, int, int], None] = None):
        if self.type == "photo":
            return self._client.http.download_media(self.direct_url, filename, progress_callback)
        elif self.type == "video":
            _res = [eval(stream.res) for stream in self.streams if stream.res]
            max_res = max(_res)
            for stream in self.streams:
                if eval(stream.res) == max_res:
                    file_format = stream.content_type.split("/")[-1]
                    if not file_format == "x-mpegURL":
                        return self._client.http.download_media(stream.url, filename, progress_callback)
        elif self.type == "animated_gif":
            file_format = self.streams[0].content_type.split("/")[-1]
            if not file_format == "x-mpegURL":
                return self._client.http.download_media(self.streams[0].url, filename, progress_callback)
        return None


class Stream(dict):
    def __init__(self, videoDict, length, ratio, client):
        super().__init__()
        self._raw = videoDict
        self._client = client
        self.bitrate = self._raw.get("bitrate")
        self.content_type = self._raw.get("content_type")
        self.url = self.direct_url = self._raw.get("url")
        self.length = length
        self.aspect_ratio = ratio
        self.res = self._get_resolution()

        for k, v in vars(self).items():
            if not k.startswith("_"):
                self[k] = v

    def _get_resolution(self):
        result = re.findall("/(\d+)x(\d+)/", self.url)
        if result:
            result = result[0]
            return f"{result[0]}*{result[1]}"

        return None

    def __repr__(self):
        return f"Stream(content_type={self.content_type}, length={self.length}, bitrate={self.bitrate}, res={self.res})"

    def download(self, filename: str = None, progress_callback: Callable[[str, int, int], None] = None):
        return self._client.http.download_media(self.url, filename, progress_callback)


class MediaSize(dict):
    def __init__(self, name, data):
        super().__init__()
        self._json = data
        self.name = self['name'] = name
        self.width = self['width'] = self._json.get('w')
        self.height = self['height'] = self._json.get('h')
        self.resize = self['resize'] = self._json.get('resize')

    def __repr__(self):
        return "MediaSize(name={}, width={}, height={}, resize={})".format(
            self.name, self.width, self.height, self.resize
        )


class ShortUser(dict):
    def __init__(self, user_dict):
        super().__init__()
        self.__raw = user_dict
        self._indices = self.__raw.get("indices")
        self.id = self.__raw.get("id_str")
        self.name = self.__raw.get("name")
        self.screen_name = self.username = self.__raw.get("screen_name")
        self.url = "https://twitter.com/{}".format(self.username)
        self.from_index = self._indices[0] if self._indices else None
        self.to_index = self._indices[1] if self._indices else None

        for k, v in vars(self).items():
            if not k.startswith("_"):
                self[k] = v

    def __eq__(self, other):
        if isinstance(other, ShortUser):
            return self.id == other.id
        elif isinstance(other, (int, str)):
            return str(self.id) == str(other)

        return False

    def __repr__(self):
        return f"ShortUser(id={self.id}, name={self.name})"


class Trends:
    def __init__(self, trend_item):
        self._raw = trend_item
        self._trend = find_objects(self._raw, "trend", None)
        self.name = self._get_name()
        self.url = self._get_url()
        self.tweet_count = self._get_count()

    def _get_name(self):
        return self._trend.get('name')

    def _get_url(self):
        url = self._trend['url'].get('url')
        url = url.replace("twitter://", "https://twitter.com/").replace("query", "q")
        return url

    def _get_count(self):
        return self._trend.get('trendMetadata', {}).get('metaDescription')

    def __repr__(self):
        return f"Trends(name={self.name})"


class Card(dict):
    def __init__(self, card_dict):
        super().__init__()
        self._dict = card_dict
        self._bindings = self._dict['legacy'].get("binding_values")
        self.rest_id = self._dict.get("rest_id")
        self.name = self._dict['legacy'].get("name")
        self.choices = []
        self.end_time = None
        self.last_updated_time = None
        self.duration = None
        self.user_ref = [User(user, None) for user in self._dict['legacy']["user_refs"]] if self._dict['legacy'].get(
            "user_refs") else []
        self.__parse_choices()

        for k, v in vars(self).items():
            if not k.startswith("_"):
                self[k] = v

    def __parse_choices(self):
        for _ in self._bindings:
            _key = _.get("key").split("_")
            if "choice" in _key[0] and "label" in _key[1]:
                _cardName = _key[0]
                _cardValue = _['value']['string_value']
                _cardValueType = _['value']['type']
                _cardCounts = 0
                _cardCountsType = None
                for __ in self._bindings:
                    __key = __.get("key")
                    if __key[0] == _key[0] and "count" in __key[1]:
                        _cardCounts = __['value']['string_value']
                        _cardCountsType = __['value']['type']
                _r = {
                    "card_name": _cardName,
                    "card_value": _cardValue,
                    "card_value_type": _cardValueType,
                    "card_counts": _cardCounts,
                    "card_counts_type": _cardCountsType,
                }
                self.choices.append(Choice(_r))
            elif _key[0] == "end" and _key[1] == "datetime":
                self.end_time = parser.parse(_['value']['string_value'])
                # last_updated_datetime_utc
            elif _key[0] == "last" and _key[1] == "updated":
                self.last_updated_time = parser.parse(_['value']['string_value'])
                # duration_minutes
            elif _key[0] == "duration" and _key[1] == "minutes":
                self.duration = _['value']['string_value']

    def __repr__(self):
        return f"Card(id={self.rest_id}, choices={len(self.choices) if self.choices else []}, end_time={self.end_time}, duration={len(self.duration) if self.duration else 0} minutes)"

class Poll(dict):
    CHOICE_LABEL_REGEX = r"choice\d+_label"
    CHOICE_COUNT_REGEX = r"choice\d+_count"
    CHOICE_COUNT_FORMAT = "choice{}_count"

    def __init__(self, card):
        super().__init__()
        self._raw = card
        self._pool = self._raw['legacy']
        self._parsed = self._parse_keys()
        self.id = self._get_id()
        self.name = self._get_name()
        self.choices = self._get_choices()
        self.end_time = self._get_end_time()
        self.last_updated_time = self._get_last_updated_time()
        self.is_final = self._get_is_final()
        self.duration = self._get_duration()
        self.selected_choice = self._get_selected_choice()
        self.user_ref = self._get_user_ref()

        for k, v in vars(self).items():
            if not k.startswith("_"):
                self[k] = v

    def __repr__(self):
        return "Pool(id={}, end_time={}, duration={} minutes, is_final={}, choices={})".format(
            self.id, self.end_time, self.duration, self.is_final, self.choices
        )

    def _parse_keys(self):
        parsed = {}
        if isinstance(self._pool['binding_values'], list):
            for value in self._pool['binding_values']:
                key, value_ = value['key'], value['value']
                parsed[key] = value_
        elif isinstance(self._pool['binding_values'], dict):
            parsed = self._pool['binding_values']
        return parsed

    def _get_name(self):
        return self._pool.get('name')

    def _get_id(self):
        return self._raw.get('rest_id') or self._pool.get('url', '').replace("\\", "")

    def _get_choices(self):
        results = []

        for key, value in self._parsed.items():
            if re.match(self.CHOICE_LABEL_REGEX, key):
                number = re.findall("\d+", key)[0]
                string_value = value['string_value']
                choice_count = self._parsed.get(self.CHOICE_COUNT_FORMAT.format(number), {}).get('string_value', "0")

                results.append(Choice(number, self.id, self.name, string_value, choice_count))
        return results

    def _get_selected_choice(self):
        choice = self._parsed.get('selected_choice', None)

        if not choice:
            return None

        for _choice in self.choices:
            if _choice.key == choice['string_value']:
                return _choice

        return None


    def _get_user_ref(self):
        return [User(user, None) for user in self._pool.get('user_refs_results', [])]

    def _get_end_time(self):
        return parse_time(self._parsed.get('end_datetime_utc', {}).get('string_value'))

    def _get_last_updated_time(self):
        return parse_time(self._parsed.get('last_updated_datetime_utc', {}).get('string_value'))

    def _get_duration(self):
        return self._parsed.get('duration_minutes', {}).get('string_value', "0")

    def _get_is_final(self):
        return self._parsed.get('counts_are_final', {}).get('boolean_value', False)


class Choice(dict):
    def __init__(self, key, pool_id, pool_name, choice_value, choice_count):
        super().__init__()
        self.key = key
        self.name = pool_name
        self.id = pool_id
        self.value = choice_value
        self.counts = choice_count

        for k, v in vars(self).items():
            if not k.startswith("_"):
                self[k] = v

    def __repr__(self):
        return f"Choice(key={self.key}, value={self.value}, counts={self.counts})"


class Place(dict):
    def __init__(self, place_dict):
        super().__init__()
        self.__dict = place_dict
        self.id = self.__dict.get("id")
        self.country = self.__dict.get("country")
        self.country_code = self.__dict.get("country_code")
        self.full_name = self.__dict.get("full_name")
        self.name = self.__dict.get("name")
        self.url = self.__dict.get("url")
        self.coordinates = self.parse_coordinates()

        for k, v in vars(self).items():
            if not k.startswith("_"):
                self[k] = v

    def parse_coordinates(self):
        results = []
        if not self.__dict.get("bounding_box"):
            return results

        for i in self.__dict['bounding_box'].get("coordinates"):
            for p in i:
                coordinates = [p[1], p[0]]
                if coordinates not in results:
                    results.append([coordinates[0], coordinates[1]])

        return [Coordinates(i[0], i[1]) for i in results]

    def __repr__(self):
        return f"Place(id={self.id}, name={self.name}, country={self.country, self.country_code}, coordinates={self.coordinates})"


class Coordinates(dict):
    def __init__(self, latitude, longitude):
        super().__init__()
        self.latitude = latitude
        self.longitude = longitude

        for k, v in vars(self).items():
            if not k.startswith("_"):
                self[k] = v

    def __repr__(self):
        return f"Coordinates(latitude={self.latitude}, longitude={self.longitude})"


class User(dict):
    def __init__(self, user_data, client, *args):
        super().__init__()
        self._raw = user_data
        self._client = client
        self._user = find_objects(self._raw, "__typename", "User", recursive=False)

        if not self._user:
            if find_objects(self._raw, "__typename", "UserUnavailable", recursive=False):
                raise UserProtected(response=user_data)
            else:
                raise UserNotFound(response=user_data)

        self.original_user = self._user['legacy'] if self._user.get('legacy') else self._user
        self.id = self.rest_id = self.get_id()
        self.created_at = self.date = self.get_created_at()
        self.entities = self._get_key("entities")
        self.description = self.bio = self._get_key("description")
        self.fast_followers_count = self._get_key("fast_followers_count", default=0)
        self.favourites_count = self._get_key("favourites_count", default=0)
        self.followers_count = self._get_key("followers_count", default=0)
        self.friends_count = self._get_key("friends_count", default=0)
        self.has_custom_timelines = self._get_key("has_custom_timelines", default=False)
        self.is_translator = self._get_key("is_translator", default=False)
        self.listed_count = self._get_key("listed_count", default=0)
        self.location = self._get_key("location")
        self.media_count = self._get_key("media_count", default=0)
        self.name = self._get_key("name")
        self.normal_followers_count = self._get_key("normal_followers_count", default=0)
        self.profile_banner_url = self._get_key("profile_banner_url")
        self.profile_image_url_https = self._get_key("profile_image_url_https")
        self.profile_interstitial_type = self._get_key("profile_interstitial_type")
        self.protected = self._get_key("protected", default=False)
        self.screen_name = self.username = self._get_key("screen_name")
        self.statuses_count = self._get_key("statuses_count", default=0)
        self.translator_type = self._get_key("translator_type")
        self.verified = self._get_verified()
        self.can_dm = self._get_key("can_dm")
        self.following = self._get_key("following", False)
        self.community_role = self._get_key("community_role", None)
        self.notifications_enabled = self.notifications = self._get_key("notifications", False)
        # self.verified_type = self._get_key("verified_type")
        self.possibly_sensitive = self._get_key("possibly_sensitive", default=False)
        self.pinned_tweets = self._get_key("pinned_tweet_ids_str")
        self.profile_url = "https://twitter.com/{}".format(self.screen_name)

        for k, v in vars(self).items():
            if not k.startswith("_"):
                self[k] = v

    def __eq__(self, other):
        if isinstance(other, User):
            return self.id == other.id
        elif isinstance(other, (int, str)):
            return str(self.id) == str(other)

        return False

    def __repr__(self):
        return "User(id={}, username={}, name={}, verified={})".format(
            self.id, self.username, self.name, self.verified
        )

    def follow(self):
        return self._client.follow_user(self.id)

    def unfollow(self):
        return self._client.unfollow_user(self.id)

    def enable_notifications(self):
        if not self.notifications:
            return self._client.enable_user_notification(self.id)

        return True

    def disable_notifications(self):
        if self.notifications:
            return self._client.disable_user_notification(self.id)

        return True

    def _get_verified(self):
        verified = self._get_key("verified", False)
        if verified is False:
            verified = self._get_key("is_blue_verified", False)

        if verified is False:
            verified = self._get_key("ext_is_blue_verified", False)

        return False if verified in (None, False) else True

    def get_id(self):
        raw_id = self._user.get("id")

        if not str(raw_id).isdigit():
            raw_id = decodeBase64(raw_id).split(":")[-1]

        return int(raw_id)

    def get_created_at(self):
        return parse_time(self.original_user.get('created_at'))

    def _get_key(self, key, default=None):
        keyValue = default

        if self._user.get(key):
            keyValue = self._user[key]

        if self.original_user.get(key):
            keyValue = self.original_user[key]

        if str(keyValue).isdigit():
            keyValue = int(keyValue)

        return keyValue


class PeriScopeUser(dict):
    def __init__(self, user_data):
        super().__init__()
        self._raw = user_data
        self.id = self._raw.get('periscope_user_id')
        self.twitter_screen_name = self.username = self._raw.get('twitter_screen_name')
        self.display_name = self.name = self._raw.get('display_name')
        self.is_verified = self._raw.get('is_verified')
        self.twitter_id = self._raw['user_results'].get('rest_id')


class AudioSpace(dict):
    def __init__(self, audio_space, client):
        super().__init__()
        self._raw = audio_space
        self._client = client
        self._space = find_objects(self._raw, "audioSpace", None, recursive=False)
        self._meta_data = find_objects(self._space, "metadata", None, recursive=False)
        self._participants = find_objects(self._raw, "participants", None, recursive=False)
        self.id = self._meta_data.get('rest_id')
        self.state = self._meta_data.get('state')
        self.title = self._meta_data.get('title')
        self.media_key = self._meta_data.get('media_key')
        self.created_at = self.ts_to_datetime(self._meta_data.get('created_at'))
        self.started_at = self.ts_to_datetime(self._meta_data.get('started_at'))
        self.ended_at = self.ts_to_datetime(self._meta_data.get('ended_at'))
        self.updated_at = self.ts_to_datetime(self._meta_data.get('updated_at'))
        self.creator = User(self._meta_data.get('creator_results'), self._client)
        self.total_live_listeners = self._meta_data.get('total_live_listeners')
        self.total_replay_watched = self._meta_data.get('total_replay_watched')
        self.disallow_join = self._meta_data.get('disallow_join')
        self.is_employee_only = self._meta_data.get('is_employee_only')
        self.is_locked = self._meta_data.get('is_locked')
        self.is_muted = self._meta_data.get('is_muted')
        self.tweet = Tweet(self._meta_data.get('tweet_results'), self._client)
        self.admins = self._get_participants('admins')
        self.speakers = self._get_participants('speakers')

    def _get_participants(self, participant):
        return [PeriScopeUser(user) for user in self._participants[participant]]

    def get_stream_link(self):
        return self._client.http.get_audio_stream(self.media_key)

    @staticmethod
    def ts_to_datetime(ts):
        try:
            return datetime.datetime.fromtimestamp(int(ts))
        except ValueError:
            return datetime.datetime.fromtimestamp(int(ts) / 1000)

    def __repr__(self):
        return "AudioSpace(id={}, title={}, state={}, tweet={})".format(
            self.id, self.title, self.state, self.tweet
        )


class Community(dict):
    def __init__(self, data, client):
        super().__init__()
        self._raw = data
        self._client = client
        self._community = find_objects(self._raw, "__typename", "Community", recursive=False)
        self.id = self._get_id()
        self.date = self.created_at = self._get_date()
        self.description = self._get_description()
        self.name = self._get_name()
        self.role = self._get_role()
        self.member_count = self._get_member_count()
        self.moderator_count = self._get_moderator_count()
        self.admin = self._get_admin()
        self.creator = self._get_creator()
        self.rules = self._get_rules()

    def __repr__(self):
        return "Community(id={}, name={}, role={}, admin={})".format(
            self.id, self.name, self.role, self.admin
        )

    def _get_id(self):
        return self._community.get('id_str')

    def _get_date(self):
        return parse_time(self._community.get('created_at'))

    def _get_description(self):
        return self._community.get('description')

    def _get_name(self):
        return self._community.get('name')

    def _get_member_count(self):
        return self._community.get('member_count')

    def _get_moderator_count(self):
        return self._community.get('moderator_count')

    def _get_admin(self):
        return User(self._community['admin_results'], self._client)

    def _get_creator(self):
        return User(self._community['creator_results'], self._client)

    def _get_rules(self):
        return [rule['name'] for rule in self._community.get('rules', [])]

    def _get_role(self):
        return self._community.get('role')

class List(dict):
    def __init__(self, list_data, client):
        super().__init__()
        self._raw = list_data
        self._client = client
        self._list = self._get_list()
        self.id = self._get_id()
        self.name = self._get_name()
        self.created_at = self.date = self._get_date()
        self.description = self._get_description()
        self.is_member = self._get_is_member()
        self.member_count = self._get_member_count()
        self.subscriber_count = self._get_subscriber_count()
        self.admin = self._get_admin()
        self.mode = self._get_mode()

        for k, v in vars(self).items():
            if not k.startswith("_"):
                self[k] = v

    def __eq__(self, other):
        if isinstance(other, List):
            return self.id == other.id
        elif isinstance(other, (int, str)):
            return str(self.id) == str(other)

        return False

    def __repr__(self):
        return "List(id={}, name={}, admin={}, subscribers={})".format(
            self.id, self.name, self.admin, self.subscriber_count
        )

    def _get_list(self):
        if self._raw.get('list'):
            return self._raw['list']

        return self._raw

    def _get_id(self):
        if self._list.get('id_str'):
            return int(self._list['id_str'])

        if self._list.get("id"):
            _id = decodeBase64(self._list['id'])
            return int(str(_id).replace("List:",""))

        return None

    def _get_name(self):
        return self._list.get('name')

    def _get_date(self):
        return parse_time(self._list.get('created_at'))

    def _get_description(self):
        return self._list.get("description")

    def _get_is_member(self):
        return self._list.get('is_member', False)

    def _get_member_count(self):
        return self._list.get('member_count', 0)

    def _get_subscriber_count(self):
        return self._list.get('subscriber_count', 0)

    def _get_mode(self):
        return self._list.get('mode')

    def _get_admin(self):
        if not self._list.get('user_results'):
            return None

        return User(self._list['user_results'], client=self._client)